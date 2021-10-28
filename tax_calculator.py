import openpyxl as xl


def tax_calculation(list_of_files):
    n = len(list_of_files)
    for j in range(0, n):
        name = list_of_files[j]
        wb = xl.load_workbook(name)
        number_of_sheets = len(wb.sheetnames)
        for i in range(1, number_of_sheets+1):
            sheet = wb[f"Sheet{i}"]
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row, 2)
                if cell.value <= 250000:
                    tax_cell = sheet.cell(row, 3)
                    tax_cell.value = 0
                elif (cell.value > 250000) and (cell.value <= 500000):
                    tax_cell = sheet.cell(row, 3)
                    tax_cell.value = ((5*cell.value)/100)
                elif (cell.value > 500000) and (cell.value <= 1000000):
                    tax_cell = sheet.cell(row, 3)
                    tax_cell.value = ((20*cell.value)/100)
                elif cell.value > 1000000:
                    tax_cell = sheet.cell(row, 3)
                    tax_cell.value = ((30*cell.value)/100)
        wb.save(f"result{j}.xlsx")








